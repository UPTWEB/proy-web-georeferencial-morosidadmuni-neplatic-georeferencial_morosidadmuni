from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


class ReportGenerator:
    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_excel(self, filename: str, sections: dict[str, list[dict]]) -> str:
        path = self.output_dir / filename
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        header_fill = PatternFill("solid", fgColor="1a472a") # Verde institucional
        header_font = Font(color="FFFFFF", bold=True)
        alt_fill = PatternFill("solid", fgColor="f5f5f5")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(vertical="center", wrap_text=True)

        for sheet_name, rows in sections.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            if not rows:
                ws.append(["Sin datos"])
                continue
            headers = list(rows[0].keys())
            ws.append(headers)
            
            for i, row in enumerate(rows):
                ws.append([row.get(h) for h in headers])
                # Aplicar estilos a la fila
                for cell in ws[i+2]:
                    cell.border = thin_border
                    cell.alignment = align_center
                    if i % 2 != 0:
                        cell.fill = alt_fill
                        
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            
            # Autoajuste de columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50) # máximo 50 de ancho

        wb.save(path)
        return str(path)

    def export_pdf(self, filename: str, title: str, sections: dict[str, list[dict]]) -> str:
        path = self.output_dir / filename
        doc = SimpleDocTemplate(str(path), pagesize=landscape(A3), title=title, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para el texto de la tabla (más pequeño y permite word wrap)
        table_text_style = styles["BodyText"].clone("TableText")
        table_text_style.fontSize = 7
        table_text_style.leading = 8
        
        story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

        for section_name, rows in sections.items():
            story.append(Paragraph(section_name, styles["Heading2"]))
            story.append(Spacer(1, 6))
            if not rows:
                story.append(Paragraph("Sin datos disponibles.", styles["BodyText"]))
                story.append(Spacer(1, 10))
                continue
                
            headers = list(rows[0].keys())
            
            # Formatear headers
            table_data = [[Paragraph(f"<b>{h}</b>", table_text_style) for h in headers]]
            
            # Formatear celdas con Paragraph para evitar desbordamiento
            for row in rows:
                table_data.append([Paragraph(str(row.get(h, "")), table_text_style) for h in headers])
            
            # Calcular anchos proporcionales
            col_max_len = []
            for h in headers:
                max_l = len(str(h))
                for row in rows:
                    l = len(str(row.get(h, "")))
                    if l > max_l:
                        max_l = l
                col_max_len.append(min(max_l, 60)) # Cap at 60 chars equivalent
                
            total_len = sum(col_max_len)
            available_width = 1190 - 60 # A3 landscape width is ~1190 points
            col_widths = [(l / total_len) * available_width for l in col_max_len] if total_len else None
            
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a472a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ]))
            story.append(table)
            story.append(Spacer(1, 14))

        doc.build(story)
        return str(path)

    @staticmethod
    def timestamped_name(prefix: str, extension: str) -> str:
        return f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extension.lstrip('.')}"

